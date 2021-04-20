from coordinador import app, db, bcrypt
from coordinador.forms import RegistrationForm, LoginForm, SubscribeForm
from coordinador.models import User, Subscription
from flask import render_template, jsonify, redirect, url_for, request
from flask_login import login_user, current_user, logout_user
from datetime import date, timedelta
import requests, zipfile, os
from openpyxl import load_workbook
import sendgrid
from sendgrid.helpers.mail import *

def set_background(data, subscriptions):
    mydata = data
    for s in subscriptions:
        for i in range(1,25):
            print(i)
            check = mydata[s.name][str(i)]
            if check > s.trigger:
                mydata[s.name][str(i)+'style']="border: 1px solid black; padding: 5px; background-color: #F5FC0E"
            else:
                mydata[s.name][str(i)+'style']="border: 1px solid black; padding: 5px"
        mydata[s.name]['Totalstyle']="border: 1px solid black; padding: 5px"
    return mydata

# get file url
def get_url():
    myDate = date.today() + timedelta(days=1)
    year = myDate.year
    month1 = '0{month}'.format(month=myDate.month) if myDate.month < 10 else myDate.month
    if int(myDate.day) == 1:
        month1 = '0{month}'.format(month=myDate.month - 1) if myDate.month - 1 < 10 else myDate.month - 1
    month2 = '0{month}'.format(month=myDate.month) if myDate.month < 10 else myDate.month
    day = '0{day}'.format(day=myDate.day) if myDate.day < 10 else myDate.day
    return 'https://www.coordinador.cl/wp-content/uploads/{year}/{month1}/PROGRAMA{year}{month2}{day}.zip'.format(year = year, month1 = month1, month2 = month2, day = day)

# download zip file from coordinators site
def download_url(url, save_path, chunk_size=128):
    r = requests.get(url, stream=True)
    with open(save_path, 'wb') as fd:
        for chunk in r.iter_content(chunk_size=chunk_size):
            fd.write(chunk)

# extact content from zip file
def unzip_file(path):
    with zipfile.ZipFile(path, 'r') as zip_ref:
        zip_ref.extractall('files')

# remove unnecessary files
def remove_files(path):
    for f in os.listdir(path):
        os.remove(os.path.join(path, f))

# gets data from excel
def extract_data(path):
    subscriptions = {s.name for s in Subscription.query.all()}
    ps = load_workbook(path)
    sheet = ps['PROGRAMA']
    data = {}
    for row in range(2, sheet.max_row + 1):
        if str(sheet['C'+ str(row)].value) in subscriptions:
            print('ROW: ', row)
            print('VALUE: ', str(sheet['C'+ str(row)].value))
            data[str(sheet['C'+ str(row)].value)] = {
                '1': round(float(sheet['E'+ str(row)].value),1),
                '2': round(float(sheet['F'+ str(row)].value),1),
                '3': round(float(sheet['G'+ str(row)].value),1),
                '4': round(float(sheet['H'+ str(row)].value),1),
                '5': round(float(sheet['I'+ str(row)].value),1),
                '6': round(float(sheet['J'+ str(row)].value),1),
                '7': round(float(sheet['K'+ str(row)].value),1),
                '8': round(float(sheet['L'+ str(row)].value),1),
                '9': round(float(sheet['M'+ str(row)].value),1),
                '10': round(float(sheet['N'+ str(row)].value),1),
                '11': round(float(sheet['O'+ str(row)].value),1),
                '12': round(float(sheet['P'+ str(row)].value),1),
                '13': round(float(sheet['Q'+ str(row)].value),1),
                '14': round(float(sheet['R'+ str(row)].value),1),
                '15': round(float(sheet['S'+ str(row)].value),1),
                '16': round(float(sheet['T'+ str(row)].value),1),
                '17': round(float(sheet['U'+ str(row)].value),1),
                '18': round(float(sheet['V'+ str(row)].value),1),
                '19': round(float(sheet['W'+ str(row)].value),1),
                '20': round(float(sheet['X'+ str(row)].value),1),
                '21': round(float(sheet['Y'+ str(row)].value),1),
                '22': round(float(sheet['Z'+ str(row)].value),1),
                '23': round(float(sheet['AA'+ str(row)].value),1),
                '24': round(float(sheet['AB'+ str(row)].value),1),
                'Total': round(float(sheet['AC'+ str(row)].value),2),
                }
    return data

# creates registration form errors dictionary
def get_errors(form):
    errors = {}
    if form.username.errors:
        errors['username'] = form.username.errors
    if form.email.errors:
        errors['email'] = form.email.errors
    if form.confirm_password.errors:
        errors['confirm_password'] = form.confirm_password.errors
    return errors

@app.route('/')
def hello_world():
    if current_user.is_authenticated:
        return redirect(url_for('subscribe'))
    return render_template('landing.html')


@app.route('/get_report')
def get_report():
    sg = sendgrid.SendGridAPIClient(api_key='API_KEY')
    from_email = Email("notificadorelectrico@gmail.com")
    subject = "[Coordinador Electrico]"
    url = get_url()
    download_url(url,"./files/coordinador.zip")
    unzip_file('./files/coordinador.zip')
    myDate = date.today()  + timedelta(days=1)
    year = myDate.year
    month = '0{month}'.format(month=myDate.month) if myDate.month < 10 else myDate.month
    day = '0{day}'.format(day=myDate.day) if myDate.day < 10 else myDate.day
    myDateString = '{day}-{month}-{year}'.format(day=day, month=month, year=year)
    path = './files/PRG{year}{month}{day}.xlsx'.format(year = str(year)[2:], month = month, day = day)
    data = extract_data(path)
    if current_user.is_authenticated:
        subscriptions = current_user.subscriptions
        if subscriptions and len(subscriptions) > 0:
            to_email = To(current_user.email)
            content = Content("text/html", render_template('mail.html', user=current_user, data=set_background(data, subscriptions), subscriptions=subscriptions, date=myDateString))
            mail = Mail(from_email, to_email, subject, content)
            response = sg.client.mail.send.post(request_body=mail.get())
            return redirect(url_for('subscribe'))
    else:
        users = User.query.all()
        for u in users:
            subscriptions = u.subscriptions
            if subscriptions and len(subscriptions) > 0:
                to_email = To(u.email)
                content = Content("text/html", render_template('mail.html', user=u, data=set_background(data, subscriptions), subscriptions=subscriptions, date=myDateString))
                mail = Mail(from_email, to_email, subject, content)
                response = sg.client.mail.send.post(request_body=mail.get())
        remove_files('./files')
        return 'about router'

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('subscribe'))
    form = RegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(username=form.username.data, email=form.email.data, password=hashed_password)
        db.session.add(user)
        db.session.commit()
        login_user(user, remember=True)
        return redirect(url_for('subscribe'))
    return render_template('register.html', form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('subscribe'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user, remember=True)
            response = { 'status': 200, 'currentUser': str(current_user) }
            return redirect(url_for('subscribe'))
        return jsonify({ 'status': 400, 'error': 'Login Unsuccessful. Please check Email and Password' })
    return render_template('login.html', form=form)

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('hello_world'))

@app.route('/subscribe', methods=['GET', 'POST'])
def subscribe():
    if current_user.is_authenticated:
        form = SubscribeForm()
        if form.validate_on_submit():
            subscription = Subscription(name=form.name.data, user_id=current_user.id, trigger=form.trigger.data)
            db.session.add(subscription)
            db.session.commit()
        if form.name.errors:
            print(form.name.errors)
        if form.trigger.errors:
            print(form.trigger.errors)
        subscriptions = current_user.subscriptions
        return render_template('subscribe.html', form=form, subscriptions=subscriptions)
    return jsonify({ 'error': 'user not authenticated'})

@app.route('/unsubscribe', methods=['POST'])
def delete_subscription():
    if current_user.is_authenticated:
        subscription_id = request.form.get("subscription")
        subscription = Subscription.query.filter_by(id=int(subscription_id)).first()
        if current_user.id == subscription.user_id:
            db.session.delete(subscription)
            db.session.commit()
        subscriptions = current_user.subscriptions
        return render_template('subscribe.html', form=form, subscriptions=subscriptions)
    else:
        return jsonify({ 'error': 'user not authenticated'})

@app.route("/submit_update", methods=["POST"])
def update_subscription():
    if current_user.is_authenticated:
        subscription_id = request.form.get("subscription")
        subscription_name = request.form.get("name")
        subscription_trigger = request.form.get("trigger")
        subscription = Subscription.query.filter_by(id=subscription_id).first()
        subscription.name = subscription_name
        subscription.trigger = subscription_trigger
        if current_user.id == subscription.user_id:
            db.session.commit()
        subscriptions = current_user.subscriptions
        form = SubscribeForm()
        return render_template('subscribe.html', form=form, subscriptions=subscriptions)
    else:
        return jsonify({ 'error': 'user not authenticated'})
